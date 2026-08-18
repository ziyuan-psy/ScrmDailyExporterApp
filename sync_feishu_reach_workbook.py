#!/usr/bin/env python3
"""Append local reach workbook rows to a Feishu spreadsheet.

The sync is intentionally append-only:
- it reads the latest date already present in each online sheet
- it appends local rows only when the target date is newer than that sheet
- it never clears or overwrites existing online rows
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel, to_excel

import export_reach_daily_excel as reach_excel


BASE_URL = "https://open.feishu.cn/open-apis"
DEFAULT_SCAN_ROWS = 5000
SUMMARY_DATA_START_ROW = 3
STORE_GROUP_DATA_START_ROW = 2
SUMMARY_COLUMNS = 9
STORE_GROUP_COLUMNS = 5
ALIGN_CENTER = 1
DATE_FORMATTER = "yyyy/MM/dd"
NUMBER_FORMATTER = "#,##0"
WEEKDAYS = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]


class FeishuSyncError(RuntimeError):
    pass


@dataclass(frozen=True)
class FeishuConfig:
    enabled: bool
    app_id: str
    app_secret: str
    spreadsheet_token: str
    wiki_token: str
    summary_sheet_id: str
    store_group_sheet_id: str
    scan_rows: int


@dataclass
class OnlineSheetState:
    values: List[List[Any]]
    existing_dates: set[date]
    max_date: Optional[date]
    last_row: int


@dataclass(frozen=True)
class ParsedSheetLink:
    token_type: str
    token: str
    sheet_id: str


def load_dotenv(path: Path) -> Dict[str, str]:
    values: Dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export ") :].strip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            values[key] = value
    return values


def env_value(env_file: Dict[str, str], key: str, default: str = "") -> str:
    return os.environ.get(key) or env_file.get(key) or default


def env_bool(env_file: Dict[str, str], key: str, default: bool = False) -> bool:
    value = env_value(env_file, key, "1" if default else "0").strip().lower()
    return value not in {"", "0", "false", "no", "off", "disabled"}


def env_int(env_file: Dict[str, str], key: str, default: int) -> int:
    value = env_value(env_file, key, str(default)).strip()
    try:
        return max(1, int(value))
    except ValueError:
        return default


def parse_wiki_token(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    if value.startswith("http://") or value.startswith("https://"):
        parsed = urllib.parse.urlparse(value)
        match = re.search(r"/wiki/([^/?#]+)", parsed.path)
        return match.group(1) if match else ""
    return value


def parse_sheet_link(value: str) -> ParsedSheetLink:
    value = value.strip()
    if not value:
        raise FeishuSyncError("Feishu sheet link is empty.")
    parsed = urllib.parse.urlparse(value)
    if parsed.scheme not in {"http", "https"}:
        raise FeishuSyncError("Please paste a full Feishu sheet URL.")

    query = urllib.parse.parse_qs(parsed.query)
    sheet_id = str((query.get("sheet") or [""])[0]).strip()
    if not sheet_id:
        raise FeishuSyncError("Feishu sheet URL is missing the sheet id.")

    wiki_match = re.search(r"/wiki/([^/?#]+)", parsed.path)
    if wiki_match:
        return ParsedSheetLink("wiki", wiki_match.group(1), sheet_id)

    sheet_match = re.search(r"/sheets/([^/?#]+)", parsed.path)
    if sheet_match:
        return ParsedSheetLink("spreadsheet", sheet_match.group(1), sheet_id)

    raise FeishuSyncError("Please paste a Feishu wiki or spreadsheet URL.")


def link_config_updates(summary_url: str, store_group_url: str) -> Dict[str, str]:
    summary = parse_sheet_link(summary_url)
    store_group = parse_sheet_link(store_group_url)
    if summary.token_type != store_group.token_type or summary.token != store_group.token:
        raise FeishuSyncError("The two Feishu links must come from the same online spreadsheet.")

    updates = {
        "FEISHU_SUMMARY_SHEET_URL": summary_url.strip(),
        "FEISHU_STORE_GROUP_SHEET_URL": store_group_url.strip(),
        "FEISHU_SUMMARY_SHEET_ID": summary.sheet_id,
        "FEISHU_STORE_GROUP_SHEET_ID": store_group.sheet_id,
    }
    if summary.token_type == "wiki":
        updates["FEISHU_WIKI_TOKEN"] = summary.token
        updates["FEISHU_SPREADSHEET_TOKEN"] = ""
    else:
        updates["FEISHU_SPREADSHEET_TOKEN"] = summary.token
        updates["FEISHU_WIKI_TOKEN"] = ""
    return updates


def update_dotenv_values(path: Path, updates: Dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = path.read_text(encoding="utf-8-sig").splitlines() if path.exists() else []
    remaining = dict(updates)
    output: List[str] = []
    key_re = re.compile(r"^\s*(?:export\s+)?([A-Za-z_][A-Za-z0-9_]*)\s*=")
    for line in existing:
        match = key_re.match(line)
        if match and match.group(1) in remaining:
            key = match.group(1)
            output.append(f"{key}={remaining.pop(key)}")
        else:
            output.append(line)
    if remaining and output and output[-1].strip():
        output.append("")
    for key, value in remaining.items():
        output.append(f"{key}={value}")
    path.write_text("\n".join(output) + "\n", encoding="utf-8")


def sync_config_updates(
    app_id: str,
    app_secret: str,
    summary_url: str,
    store_group_url: str,
    enabled: bool,
) -> Dict[str, str]:
    updates = link_config_updates(summary_url, store_group_url)
    updates["FEISHU_APP_ID"] = app_id.strip()
    updates["FEISHU_APP_SECRET"] = app_secret.strip()
    updates["FEISHU_SYNC_ENABLED"] = "1" if enabled else "0"
    return updates


def save_sync_config(
    env_path: Path,
    app_id: str,
    app_secret: str,
    summary_url: str,
    store_group_url: str,
    enabled: bool,
) -> None:
    updates = sync_config_updates(app_id, app_secret, summary_url, store_group_url, enabled)
    update_dotenv_values(env_path, updates)


def save_sheet_links(env_path: Path, summary_url: str, store_group_url: str, enabled: bool) -> None:
    env_file = load_dotenv(env_path)
    save_sync_config(
        env_path,
        env_value(env_file, "FEISHU_APP_ID"),
        env_value(env_file, "FEISHU_APP_SECRET"),
        summary_url,
        store_group_url,
        enabled,
    )


def load_config(env_path: Path) -> FeishuConfig:
    env_file = load_dotenv(env_path)
    link_updates: Dict[str, str] = {}
    summary_url = env_value(env_file, "FEISHU_SUMMARY_SHEET_URL").strip()
    store_group_url = env_value(env_file, "FEISHU_STORE_GROUP_SHEET_URL").strip()
    if summary_url and store_group_url:
        link_updates = link_config_updates(summary_url, store_group_url)
    return FeishuConfig(
        enabled=env_bool(env_file, "FEISHU_SYNC_ENABLED", False),
        app_id=env_value(env_file, "FEISHU_APP_ID").strip(),
        app_secret=env_value(env_file, "FEISHU_APP_SECRET").strip(),
        spreadsheet_token=(link_updates.get("FEISHU_SPREADSHEET_TOKEN") or env_value(env_file, "FEISHU_SPREADSHEET_TOKEN")).strip(),
        wiki_token=parse_wiki_token(
            link_updates.get("FEISHU_WIKI_TOKEN")
            or env_value(env_file, "FEISHU_WIKI_TOKEN")
            or env_value(env_file, "FEISHU_WIKI_URL")
        ),
        summary_sheet_id=(link_updates.get("FEISHU_SUMMARY_SHEET_ID") or env_value(env_file, "FEISHU_SUMMARY_SHEET_ID")).strip(),
        store_group_sheet_id=(
            link_updates.get("FEISHU_STORE_GROUP_SHEET_ID")
            or env_value(env_file, "FEISHU_STORE_GROUP_SHEET_ID")
        ).strip(),
        scan_rows=env_int(env_file, "FEISHU_SCAN_ROWS", DEFAULT_SCAN_ROWS),
    )


def config_from_sheet_links(
    env_path: Path,
    summary_url: str,
    store_group_url: str,
    enabled: bool = True,
) -> FeishuConfig:
    env_file = load_dotenv(env_path)
    return config_from_values(
        env_path,
        env_value(env_file, "FEISHU_APP_ID"),
        env_value(env_file, "FEISHU_APP_SECRET"),
        summary_url,
        store_group_url,
        enabled,
    )


def config_from_values(
    env_path: Path,
    app_id: str,
    app_secret: str,
    summary_url: str,
    store_group_url: str,
    enabled: bool = True,
) -> FeishuConfig:
    env_file = load_dotenv(env_path)
    updates = link_config_updates(summary_url, store_group_url)
    return FeishuConfig(
        enabled=enabled,
        app_id=(app_id.strip() or env_value(env_file, "FEISHU_APP_ID")).strip(),
        app_secret=(app_secret.strip() or env_value(env_file, "FEISHU_APP_SECRET")).strip(),
        spreadsheet_token=updates.get("FEISHU_SPREADSHEET_TOKEN", "").strip(),
        wiki_token=updates.get("FEISHU_WIKI_TOKEN", "").strip(),
        summary_sheet_id=updates["FEISHU_SUMMARY_SHEET_ID"].strip(),
        store_group_sheet_id=updates["FEISHU_STORE_GROUP_SHEET_ID"].strip(),
        scan_rows=env_int(env_file, "FEISHU_SCAN_ROWS", DEFAULT_SCAN_ROWS),
    )


def require_config(config: FeishuConfig) -> None:
    missing: List[str] = []
    if not config.app_id:
        missing.append("FEISHU_APP_ID")
    if not config.app_secret:
        missing.append("FEISHU_APP_SECRET")
    if not config.spreadsheet_token and not config.wiki_token:
        missing.append("FEISHU_SPREADSHEET_TOKEN or FEISHU_WIKI_TOKEN")
    if not config.summary_sheet_id:
        missing.append("FEISHU_SUMMARY_SHEET_ID")
    if not config.store_group_sheet_id:
        missing.append("FEISHU_STORE_GROUP_SHEET_ID")
    if missing:
        raise FeishuSyncError("Missing Feishu sync config: " + ", ".join(missing))


def request_json(
    method: str,
    url: str,
    payload: Optional[Dict[str, Any]] = None,
    token: str = "",
) -> Dict[str, Any]:
    headers = {"Content-Type": "application/json; charset=utf-8"}
    if token:
        headers["Authorization"] = "Bearer " + token
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8") if payload is not None else None
    request = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            text = response.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        text = exc.read().decode("utf-8", errors="replace")
        try:
            body = json.loads(text)
        except json.JSONDecodeError:
            body = {"raw": text}
        raise FeishuSyncError(f"Feishu HTTP {exc.code}: {body}") from exc
    except urllib.error.URLError as exc:
        raise FeishuSyncError(f"Feishu network error: {exc.reason}") from exc

    try:
        data_obj = json.loads(text) if text else {}
    except json.JSONDecodeError as exc:
        raise FeishuSyncError(f"Non-JSON Feishu response: {text[:300]}") from exc
    code = data_obj.get("code")
    if code not in (None, 0):
        message = data_obj.get("msg") or data_obj.get("message") or data_obj
        raise FeishuSyncError(f"Feishu API error code={code}, message={message}")
    return data_obj


class FeishuClient:
    def __init__(self, config: FeishuConfig) -> None:
        self.config = config
        self.tenant_access_token = ""
        self.spreadsheet_token = ""

    def connect(self) -> None:
        require_config(self.config)
        token_resp = request_json(
            "POST",
            f"{BASE_URL}/auth/v3/tenant_access_token/internal",
            {"app_id": self.config.app_id, "app_secret": self.config.app_secret},
        )
        token = str(token_resp.get("tenant_access_token") or "")
        if not token:
            raise FeishuSyncError("Feishu did not return tenant_access_token.")
        self.tenant_access_token = token

        if self.config.spreadsheet_token:
            self.spreadsheet_token = self.config.spreadsheet_token
            return

        node_resp = request_json(
            "GET",
            f"{BASE_URL}/wiki/v2/spaces/get_node?"
            + urllib.parse.urlencode({"token": self.config.wiki_token}),
            token=self.tenant_access_token,
        )
        node = (node_resp.get("data") or {}).get("node") or {}
        if node.get("obj_type") != "sheet" or not node.get("obj_token"):
            raise FeishuSyncError("Feishu wiki node is not a spreadsheet.")
        self.spreadsheet_token = str(node["obj_token"])

    def read_values(self, sheet_id: str, start_col: str, end_col: str, rows: int) -> List[List[Any]]:
        range_name = f"{sheet_id}!{start_col}1:{end_col}{rows}"
        url = (
            f"{BASE_URL}/sheets/v2/spreadsheets/"
            + urllib.parse.quote(self.spreadsheet_token, safe="")
            + "/values/"
            + urllib.parse.quote(range_name, safe="")
        )
        data = request_json("GET", url, token=self.tenant_access_token)
        return ((data.get("data") or {}).get("valueRange") or {}).get("values") or []

    def append_rows(self, sheet_id: str, start_row: int, end_col: str, rows: List[List[Any]]) -> None:
        if not rows:
            return
        end_row = start_row + len(rows) - 1
        range_name = f"{sheet_id}!A{start_row}:{end_col}{end_row}"
        url = (
            f"{BASE_URL}/sheets/v2/spreadsheets/"
            + urllib.parse.quote(self.spreadsheet_token, safe="")
            + "/values_batch_update"
        )
        request_json(
            "POST",
            url,
            {"valueRanges": [{"range": range_name, "values": rows}]},
            token=self.tenant_access_token,
        )

    def set_style(self, range_name: str, style: Dict[str, Any]) -> None:
        url = (
            f"{BASE_URL}/sheets/v2/spreadsheets/"
            + urllib.parse.quote(self.spreadsheet_token, safe="")
            + "/style"
        )
        request_json(
            "PUT",
            url,
            {"appendStyle": {"range": range_name, "style": style}},
            token=self.tenant_access_token,
        )

    def format_summary_rows(self, start_row: int, row_count: int) -> None:
        if row_count <= 0:
            return
        end_row = start_row + row_count - 1
        self.set_style(
            f"{self.config.summary_sheet_id}!A{start_row}:I{end_row}",
            centered_style(),
        )
        self.set_style(
            f"{self.config.summary_sheet_id}!A{start_row}:A{end_row}",
            centered_style(DATE_FORMATTER),
        )
        for column in ("C", "F", "I"):
            self.set_style(
                f"{self.config.summary_sheet_id}!{column}{start_row}:{column}{end_row}",
                centered_style(NUMBER_FORMATTER),
            )

    def format_store_group_rows(self, start_row: int, row_count: int) -> None:
        if row_count <= 0:
            return
        end_row = start_row + row_count - 1
        self.set_style(
            f"{self.config.store_group_sheet_id}!A{start_row}:E{end_row}",
            centered_style(),
        )
        self.set_style(
            f"{self.config.store_group_sheet_id}!A{start_row}:A{end_row}",
            centered_style(DATE_FORMATTER),
        )
        self.set_style(
            f"{self.config.store_group_sheet_id}!E{start_row}:E{end_row}",
            centered_style(NUMBER_FORMATTER),
        )


def centered_style(formatter: str = "") -> Dict[str, Any]:
    style: Dict[str, Any] = {
        "hAlign": ALIGN_CENTER,
        "vAlign": ALIGN_CENTER,
        "clean": False,
    }
    if formatter:
        style["formatter"] = formatter
    return style


def is_non_empty(value: Any) -> bool:
    return value is not None and value != ""


def parse_date_value(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
        except Exception:
            return None
        if isinstance(parsed, datetime):
            return parsed.date()
        if isinstance(parsed, date):
            return parsed
        return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def date_to_serial(value: date) -> int:
    as_datetime = datetime.combine(value, datetime.min.time())
    return int(to_excel(as_datetime))


def number_or_none(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.strip() == "-":
        return None
    try:
        return int(float(str(value).replace(",", "")))
    except (TypeError, ValueError):
        return None


def pad_row(values: Sequence[Any], length: int) -> List[Any]:
    row = list(values[:length])
    while len(row) < length:
        row.append(None)
    return row


def read_local_summary_rows(workbook_path: Path) -> Dict[date, List[Any]]:
    if not workbook_path.exists():
        raise FeishuSyncError(f"Local workbook does not exist: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    if reach_excel.SHEET_REACH_SUMMARY not in wb.sheetnames:
        raise FeishuSyncError(f"Workbook missing sheet: {reach_excel.SHEET_REACH_SUMMARY}")
    ws = wb[reach_excel.SHEET_REACH_SUMMARY]
    result: Dict[date, List[Any]] = {}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        values = pad_row(raw, SUMMARY_COLUMNS)
        target_date = parse_date_value(values[0])
        if not target_date:
            continue
        result[target_date] = [
            date_to_serial(target_date),
            values[1] or WEEKDAYS[target_date.weekday()],
            number_or_none(values[2]),
            None,
            None,
            number_or_none(values[5]),
            None,
            None,
            number_or_none(values[8]),
        ]
    return result


def read_local_store_group_rows(workbook_path: Path) -> Dict[date, List[List[Any]]]:
    if not workbook_path.exists():
        raise FeishuSyncError(f"Local workbook does not exist: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    if reach_excel.SHEET_STORE_GROUP not in wb.sheetnames:
        raise FeishuSyncError(f"Workbook missing sheet: {reach_excel.SHEET_STORE_GROUP}")
    ws = wb[reach_excel.SHEET_STORE_GROUP]
    result: Dict[date, List[List[Any]]] = {}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        values = pad_row(raw, STORE_GROUP_COLUMNS)
        target_date = parse_date_value(values[0])
        if not target_date:
            continue
        result.setdefault(target_date, []).append(
            [
                date_to_serial(target_date),
                values[1],
                values[2],
                values[3],
                number_or_none(values[4]),
            ]
        )
    return result


def online_sheet_state(values: List[List[Any]], data_start_row: int) -> OnlineSheetState:
    existing_dates: set[date] = set()
    max_date: Optional[date] = None
    last_row = 0
    for index, row in enumerate(values, start=1):
        if any(is_non_empty(value) for value in row):
            last_row = index
        if index < data_start_row:
            continue
        target_date = parse_date_value(row[0] if row else None)
        if not target_date:
            continue
        existing_dates.add(target_date)
        if max_date is None or target_date > max_date:
            max_date = target_date
    return OnlineSheetState(values=values, existing_dates=existing_dates, max_date=max_date, last_row=last_row)


def read_online_states(client: FeishuClient) -> Tuple[OnlineSheetState, OnlineSheetState]:
    summary_values = client.read_values(
        client.config.summary_sheet_id,
        "A",
        "I",
        client.config.scan_rows,
    )
    store_values = client.read_values(
        client.config.store_group_sheet_id,
        "A",
        "E",
        client.config.scan_rows,
    )
    return (
        online_sheet_state(summary_values, SUMMARY_DATA_START_ROW),
        online_sheet_state(store_values, STORE_GROUP_DATA_START_ROW),
    )


def row_spans_for_dates(
    values: List[List[Any]],
    data_start_row: int,
    target_dates: set[date],
) -> List[Tuple[int, int]]:
    spans: List[Tuple[int, int]] = []
    span_start: Optional[int] = None
    previous_row: Optional[int] = None

    for index, row in enumerate(values, start=1):
        if index < data_start_row:
            continue
        row_date = parse_date_value(row[0] if row else None)
        if row_date not in target_dates:
            if span_start is not None and previous_row is not None:
                spans.append((span_start, previous_row - span_start + 1))
                span_start = None
                previous_row = None
            continue

        if span_start is None:
            span_start = index
        previous_row = index

    if span_start is not None and previous_row is not None:
        spans.append((span_start, previous_row - span_start + 1))
    return spans


def should_append(target_date: date, state: OnlineSheetState) -> bool:
    return state.max_date is None or target_date > state.max_date


def format_existing_date(env_path: Path, target_date: date) -> Dict[str, int]:
    config = load_config(env_path)
    if not config.enabled:
        print("Feishu sync disabled; set FEISHU_SYNC_ENABLED=1 to enable.")
        return {"summary": 0, "store_group": 0}

    client = FeishuClient(config)
    client.connect()
    summary_state, store_state = read_online_states(client)
    target_dates = {target_date}

    summary_rows_formatted = 0
    for start_row, row_count in row_spans_for_dates(
        summary_state.values,
        SUMMARY_DATA_START_ROW,
        target_dates,
    ):
        client.format_summary_rows(start_row, row_count)
        summary_rows_formatted += row_count

    store_rows_formatted = 0
    for start_row, row_count in row_spans_for_dates(
        store_state.values,
        STORE_GROUP_DATA_START_ROW,
        target_dates,
    ):
        client.format_store_group_rows(start_row, row_count)
        store_rows_formatted += row_count

    print(
        f"Feishu format {target_date:%Y-%m-%d}: "
        f"summary rows formatted={summary_rows_formatted}, "
        f"store-group rows formatted={store_rows_formatted}."
    )
    return {"summary": summary_rows_formatted, "store_group": store_rows_formatted}


def sync_date(workbook_path: Path, env_path: Path, target_date: date, dry_run: bool = False) -> Dict[str, int]:
    config = load_config(env_path)
    if not config.enabled:
        print("Feishu sync disabled; set FEISHU_SYNC_ENABLED=1 to enable.")
        return {"summary": 0, "store_group": 0}

    client = FeishuClient(config)
    client.connect()
    summary_state, store_state = read_online_states(client)

    summary_rows_appended = 0
    store_rows_appended = 0

    if should_append(target_date, summary_state):
        local_summary = read_local_summary_rows(workbook_path)
        row = local_summary.get(target_date)
        if not row:
            raise FeishuSyncError(f"Local summary row not found for {target_date:%Y-%m-%d}.")
        start_row = max(summary_state.last_row + 1, SUMMARY_DATA_START_ROW)
        if dry_run:
            print(f"Would append 1 summary row at {config.summary_sheet_id}!A{start_row}:I{start_row}.")
        else:
            client.append_rows(config.summary_sheet_id, start_row, "I", [row])
            client.format_summary_rows(start_row, 1)
        summary_rows_appended = 1
    else:
        print(
            "Summary sheet already has data through "
            f"{summary_state.max_date:%Y-%m-%d}; skipping {target_date:%Y-%m-%d}."
        )

    if should_append(target_date, store_state):
        local_store = read_local_store_group_rows(workbook_path)
        rows = local_store.get(target_date)
        if not rows:
            raise FeishuSyncError(f"Local store-group rows not found for {target_date:%Y-%m-%d}.")
        start_row = max(store_state.last_row + 1, STORE_GROUP_DATA_START_ROW)
        if dry_run:
            end_row = start_row + len(rows) - 1
            print(f"Would append {len(rows)} store-group rows at {config.store_group_sheet_id}!A{start_row}:E{end_row}.")
        else:
            client.append_rows(config.store_group_sheet_id, start_row, "E", rows)
            client.format_store_group_rows(start_row, len(rows))
        store_rows_appended = len(rows)
    else:
        print(
            "Store-group sheet already has data through "
            f"{store_state.max_date:%Y-%m-%d}; skipping {target_date:%Y-%m-%d}."
        )

    print(
        f"Feishu sync {target_date:%Y-%m-%d}: "
        f"summary rows appended={summary_rows_appended}, "
        f"store-group rows appended={store_rows_appended}."
    )
    return {"summary": summary_rows_appended, "store_group": store_rows_appended}


def complete_through_date(env_path: Path) -> Optional[date]:
    config = load_config(env_path)
    if not config.enabled:
        return None
    client = FeishuClient(config)
    client.connect()
    summary_state, store_state = read_online_states(client)
    if summary_state.max_date is None or store_state.max_date is None:
        return None
    return min(summary_state.max_date, store_state.max_date)


def workbook_dates_through(workbook_path: Path, cutoff: date) -> List[date]:
    summary_dates = set(read_local_summary_rows(workbook_path))
    store_dates = set(read_local_store_group_rows(workbook_path))
    return sorted((summary_dates | store_dates) & {item for item in summary_dates | store_dates if item <= cutoff})


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Append local reach workbook rows to Feishu.")
    parser.add_argument("--workbook", required=True, help="Local reach workbook path.")
    parser.add_argument("--env", required=True, help="Runtime .env path.")
    parser.add_argument("--date", required=True, help="Date to sync, YYYY-MM-DD.")
    parser.add_argument("--dry-run", action="store_true", help="Read both sides and print what would be appended.")
    parser.add_argument("--format-only", action="store_true", help="Only format existing online rows for the date.")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    target_date = datetime.strptime(args.date, "%Y-%m-%d").date()
    if args.format_only:
        format_existing_date(
            env_path=Path(args.env).expanduser(),
            target_date=target_date,
        )
        return 0
    sync_date(
        workbook_path=Path(args.workbook).expanduser(),
        env_path=Path(args.env).expanduser(),
        target_date=target_date,
        dry_run=args.dry_run,
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except FeishuSyncError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
